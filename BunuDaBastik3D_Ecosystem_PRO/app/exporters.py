from __future__ import annotations
from pathlib import Path
import openpyxl
import db


def export_table(table, rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table[:31]
    rows = list(rows)
    if not rows:
        ws.append(['Bilgi'])
        ws.append(['Kayıt bulunamadı'])
    else:
        headers = rows[0].keys()
        ws.append(list(headers))
        for r in rows:
            ws.append([r[h] for h in headers])
    for col in ws.columns:
        max_len = 10
        letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or '')))
        ws.column_dimensions[letter].width = min(max_len + 2, 45)
    out_path = Path(out_path)
    wb.save(out_path)
    return out_path


def export_all(out_path):
    wb = openpyxl.Workbook()
    first = True
    for table in ['inbox_items','email_messages','quotes','catalog','customers','orders','production_jobs','inventory','sync_queue']:
        rows = db.list_rows(table, order='id ASC')
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = table[:31]
        if rows:
            headers = rows[0].keys()
            ws.append(list(headers))
            for r in rows:
                ws.append([r[h] for h in headers])
        else:
            ws.append(['Bilgi'])
            ws.append(['Kayıt bulunamadı'])
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[letter].width = min(max_len + 2, 45)
    out_path = Path(out_path)
    wb.save(out_path)
    return out_path
